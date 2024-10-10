from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 加载Excel工作簿
path = "/LLM-Medical-Agent/data/output (4).xlsx"
wb = load_workbook(path)

# 选择第二个工作表
ws = wb.worksheets[1]  # 通过索引选择第二个工作表，索引从0开始，所以第二个表是[1]

# 定义红色和紫色的 RGB 代码
red_rgb = 'FFFF0000'  # 红色
purple_rgb = 'FF7030A0'  # 紫色

# 遍历工作表的所有单元格
for row in ws.iter_rows():
    for cell in row:
        # 获取单元格的填充色
        cell_color = cell.fill.start_color.rgb if cell.fill.start_color is not None else None

        # 判断单元格的颜色是否为红色或紫色
        if cell_color in [red_rgb, purple_rgb]:
            # 如果单元格的值是布尔类型，则进行反转
            if isinstance(cell.value, bool):
                cell.value = not cell.value  # 反转True/False
                print(f"单元格 {cell.coordinate} 的值已被反转为 {cell.value}")

# 保存修改后的Excel文件
wb.save("/LLM-Medical-Agent/data/output_modified.xlsx")